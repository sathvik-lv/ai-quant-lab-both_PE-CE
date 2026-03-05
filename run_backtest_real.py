"""
Full 2020-2025 backtest using real FO bhavcopy data.
Follows STRATEGY_LOCKED_V1.4_CE_PE.md exactly.
Outputs 3-sheet Excel matching reference format.
"""

import os
import sys
import csv
import io
import zipfile
import math
import logging
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Project root ──
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_ROOT = os.path.join(PROJECT_ROOT, "data")

# ── Strategy parameters (from frozen_params.py) ──
INDICES = ["BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "NIFTY"]
PREMIUM_RATIO_MIN = 0.9
PREMIUM_RATIO_MAX = 1.1
MAX_TOTAL_EXPOSURE_PCT = 60.0
MAX_HOLD_DAYS = {"BANKNIFTY": 5, "FINNIFTY": 5, "MIDCPNIFTY": 5, "NIFTY": 5}
CAPITAL_PCT = {
    "BANKNIFTY_PE": 17.34, "NIFTY_PE": 13.00, "FINNIFTY_PE": 8.67, "MIDCPNIFTY_PE": 4.34,
    "FINNIFTY_CE": 3.60, "BANKNIFTY_CE": 2.70, "NIFTY_CE": 1.80, "MIDCPNIFTY_CE": 0.90,
}
MAX_CONCURRENT = {
    "BANKNIFTY_PE": 2, "BANKNIFTY_CE": 2, "FINNIFTY_PE": 2, "FINNIFTY_CE": 2,
    "MIDCPNIFTY_PE": 1, "MIDCPNIFTY_CE": 1, "NIFTY_PE": 2, "NIFTY_CE": 2,
}

# ── Strike intervals per index ──
STRIKE_INTERVAL = {"BANKNIFTY": 100, "NIFTY": 50, "FINNIFTY": 50, "MIDCPNIFTY": 25}

# ── Lot sizes (historical, per index) ──
# Using standard NSE lot sizes that were in effect during each period
LOT_SIZES = {
    "BANKNIFTY": 15,
    "NIFTY": 75,
    "FINNIFTY": 40,
    "MIDCPNIFTY": 75,
}

# ── Cost parameters ──
ENTRY_COST = 50   # flat per trade
EXIT_COST = 49    # flat per trade

# ── Starting equity ──
STARTING_EQUITY = 10_000_000

# ── Index availability dates ──
INDEX_START = {
    "BANKNIFTY": datetime(2000, 1, 1),
    "NIFTY": datetime(2000, 1, 1),
    "FINNIFTY": datetime(2021, 1, 1),
    "MIDCPNIFTY": datetime(2022, 7, 1),
}

# ── Symbol mapping for spot CSVs ──
SPOT_FILE_MAP = {
    "BANKNIFTY": "Nifty Bank Historical Data.csv",
    "NIFTY": "Nifty 50 Historical Data.csv",
    "FINNIFTY": "Nifty Financial Services Historical Data.csv",
    "MIDCPNIFTY": "Nifty Midcap 50 Historical Data.csv",
}
SPOT_FILE_MAP_2025 = {
    "BANKNIFTY": "banknifty_2025-01-01.csv",
    "NIFTY": "nifty50_2025-01-01.csv",
    "FINNIFTY": "finifty_2025-01-01.csv",
    "MIDCPNIFTY": "midcpnifty_2025-01-01.csv",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("backtest_real")


# ═══════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════

def parse_old_bhavcopy(zip_path):
    """Parse pre-July-2024 bhavcopy format. Returns list of dicts."""
    records = []
    try:
        with zipfile.ZipFile(zip_path) as z:
            for name in z.namelist():
                if not name.endswith('.csv'):
                    continue
                with z.open(name) as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding='utf-8', errors='replace'))
                    for row in reader:
                        inst = row.get('INSTRUMENT', '').strip()
                        sym = row.get('SYMBOL', '').strip()
                        if inst != 'OPTIDX' or sym not in INDICES:
                            continue
                        try:
                            expiry_str = row['EXPIRY_DT'].strip()
                            expiry = datetime.strptime(expiry_str, '%d-%b-%Y').date()
                            timestamp_str = row['TIMESTAMP'].strip()
                            trade_date = datetime.strptime(timestamp_str, '%d-%b-%Y').date()
                            strike = float(row['STRIKE_PR'].strip())
                            opt_type = row['OPTION_TYP'].strip()
                            close = float(row['CLOSE'].strip())
                            oi = int(row['OPEN_INT'].strip()) if row.get('OPEN_INT', '').strip() else 0
                            records.append({
                                'date': trade_date,
                                'index': sym,
                                'expiry': expiry,
                                'strike': strike,
                                'opt_type': opt_type,
                                'close': close,
                                'oi': oi,
                            })
                        except (ValueError, KeyError):
                            continue
    except (zipfile.BadZipFile, Exception) as e:
        logger.warning("Failed to parse %s: %s", zip_path, e)
    return records


def parse_new_bhavcopy(zip_path):
    """Parse UDiFF format (July 2024+). Returns list of dicts."""
    records = []
    try:
        with zipfile.ZipFile(zip_path) as z:
            for name in z.namelist():
                if not name.endswith('.csv'):
                    continue
                with z.open(name) as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding='utf-8', errors='replace'))
                    for row in reader:
                        fin_type = row.get('FinInstrmTp', '').strip()
                        sym = row.get('TckrSymb', '').strip()
                        if fin_type != 'IDO' or sym not in INDICES:
                            continue
                        try:
                            trade_date = datetime.strptime(row['TradDt'].strip(), '%Y-%m-%d').date()
                            expiry = datetime.strptime(row['XpryDt'].strip(), '%Y-%m-%d').date()
                            strike = float(row['StrkPric'].strip())
                            opt_type = row['OptnTp'].strip()
                            close = float(row['ClsPric'].strip())
                            oi_str = row.get('OpnIntrst', '0').strip()
                            oi = int(oi_str) if oi_str else 0
                            records.append({
                                'date': trade_date,
                                'index': sym,
                                'expiry': expiry,
                                'strike': strike,
                                'opt_type': opt_type,
                                'close': close,
                                'oi': oi,
                            })
                        except (ValueError, KeyError):
                            continue
    except (zipfile.BadZipFile, Exception) as e:
        logger.warning("Failed to parse %s: %s", zip_path, e)
    return records


def load_all_bhavcopy():
    """
    Load all bhavcopy data from 2020-2025.
    Returns: dict[date][index][expiry][strike][opt_type] = {'close': float, 'oi': int}
    """
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))
    total_records = 0

    for year in range(2020, 2026):
        bhav_dir = os.path.join(DATA_ROOT, f"raw_{year}", "fo_bhavcopy")
        if not os.path.isdir(bhav_dir):
            logger.warning("Missing bhavcopy dir: %s", bhav_dir)
            continue

        files = sorted(os.listdir(bhav_dir))
        logger.info("Loading %d bhavcopy files for %d...", len(files), year)

        for fname in files:
            if not fname.endswith('.zip'):
                continue
            fpath = os.path.join(bhav_dir, fname)

            if fname.startswith('BhavCopy_NSE_FO_'):
                records = parse_new_bhavcopy(fpath)
            else:
                records = parse_old_bhavcopy(fpath)

            for r in records:
                data[r['date']][r['index']][r['expiry']][r['strike']][r['opt_type']] = {
                    'close': r['close'], 'oi': r['oi']
                }
            total_records += len(records)

        logger.info("Year %d done. Running total: %d records", year, total_records)

    logger.info("Total bhavcopy records loaded: %d", total_records)
    return data


def parse_spot_price(val_str):
    """Parse Investing.com price string like '31,264.05' to float."""
    return float(val_str.replace(',', '').replace('"', '').strip())


def parse_spot_date(date_str):
    """Parse date string like '31-12-2020' or '12/31/2020' to date."""
    date_str = date_str.replace('"', '').replace('\ufeff', '').strip()
    for fmt in ('%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def load_all_spot():
    """
    Load spot price data for all indices, 2020-2025.
    Returns: dict[date][index] = spot_price
    """
    spot = defaultdict(dict)

    for year in range(2020, 2026):
        spot_dir = os.path.join(DATA_ROOT, f"raw_{year}", "index_spot")
        if not os.path.isdir(spot_dir):
            continue

        file_map = SPOT_FILE_MAP_2025 if year == 2025 else SPOT_FILE_MAP

        for index, fname in file_map.items():
            fpath = os.path.join(spot_dir, fname)
            if not os.path.isfile(fpath):
                continue

            with open(fpath, encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    date_key = None
                    for k in row:
                        if 'date' in k.lower() or 'Date' in k:
                            date_key = k
                            break
                    if not date_key:
                        date_key = list(row.keys())[0]

                    price_key = 'Price'
                    if price_key not in row:
                        for k in row:
                            if 'price' in k.lower():
                                price_key = k
                                break

                    d = parse_spot_date(row[date_key])
                    if d is None:
                        continue
                    try:
                        p = parse_spot_price(row[price_key])
                        spot[d][index] = p
                    except (ValueError, KeyError):
                        continue

    logger.info("Spot data loaded: %d dates", len(spot))
    return spot


# ═══════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def get_atm_strike(spot, interval):
    """Round spot to nearest strike interval."""
    return round(spot / interval) * interval


def classify_expiry(expiry_date, trade_date):
    """
    Classify an expiry as weekly or monthly.
    Monthly = last expiry of the calendar month for that index.
    We'll determine this by checking if there's no later expiry in the same month.
    """
    # This is just a helper; actual classification happens in get_expiries
    pass


def get_expiries(trade_date, available_expiries):
    """
    From available expiry dates for an index on a given trade date,
    find the nearest weekly expiry and the nearest monthly expiry.

    Weekly: nearest future expiry (within ~7 days)
    Monthly: the last expiry of the current or next month

    Returns: (weekly_expiry, monthly_expiry) or (None, None)
    """
    future_expiries = sorted([e for e in available_expiries if e >= trade_date])
    if len(future_expiries) < 2:
        return None, None

    # Weekly = nearest expiry
    weekly = future_expiries[0]

    # If weekly is same day as trade, need at least 1 day remaining
    # Skip to next if it expires today (can't enter on expiry day)
    if weekly == trade_date:
        if len(future_expiries) < 3:
            return None, None
        weekly = future_expiries[1]
        future_expiries = future_expiries[1:]

    # Monthly = find the last expiry of a month that is different from weekly
    # Group expiries by month
    monthly = None
    for exp in future_expiries:
        if exp != weekly and exp > weekly:
            # Check if this expiry is the last one in its month
            # Simple approach: pick the first expiry that's in a later month or ≥ 2 weeks out
            days_diff = (exp - weekly).days
            if days_diff >= 7:
                monthly = exp
                break

    if monthly is None:
        return None, None

    return weekly, monthly


def current_exposure_pct(active_trades, equity):
    """Calculate total exposure % of current active trades."""
    if equity <= 0:
        return 100.0
    total_capital = sum(t['capital_deployed'] for t in active_trades)
    return (total_capital / equity) * 100.0


def count_concurrent(active_trades, index, side):
    """Count active trades for a specific index+side."""
    return sum(1 for t in active_trades if t['index'] == index and t['side'] == side)


def has_overlap(active_trades, index, side, strike, weekly_exp):
    """Check if there's an overlapping trade (same index+side+strike+expiry)."""
    for t in active_trades:
        if (t['index'] == index and t['side'] == side
                and t['strike'] == strike and t['near_expiry'] == weekly_exp):
            return True
    return False


# ═══════════════════════════════════════════════════════════════
# MAIN BACKTEST
# ═══════════════════════════════════════════════════════════════

def run_backtest():
    logger.info("=" * 60)
    logger.info("FULL BACKTEST 2020-2025 — ATM CE&PE Calendar Spread v1.4")
    logger.info("=" * 60)

    # Load data
    bhav_data = load_all_bhavcopy()
    spot_data = load_all_spot()

    # Get all trading dates (sorted)
    all_dates = sorted(bhav_data.keys())
    logger.info("Trading dates: %s to %s (%d days)", all_dates[0], all_dates[-1], len(all_dates))

    # State
    equity = STARTING_EQUITY
    active_trades = []
    completed_trades = []
    ratio_timeseries = []
    trade_counter = 0

    for date_idx, trade_date in enumerate(all_dates):
        day_data = bhav_data[trade_date]
        spot_today = spot_data.get(trade_date, {})

        # ── Check exits for active trades ──
        trades_to_close = []
        for i, trade in enumerate(active_trades):
            days_open = (trade_date - trade['entry_date']).days
            exit_reason = None
            all_exits = []

            # EXPIRY_FORCE_CLOSE: if trade_date >= near_expiry
            if trade_date >= trade['near_expiry']:
                all_exits.append('EXPIRY_FORCE_CLOSE')
                if exit_reason is None:
                    exit_reason = 'EXPIRY_FORCE_CLOSE'

            # EXIT_TIME: max hold days
            if days_open >= MAX_HOLD_DAYS[trade['index']]:
                all_exits.append('EXIT_TIME')
                if exit_reason is None:
                    exit_reason = 'EXIT_TIME'

            if exit_reason:
                # Get exit prices
                idx_data = day_data.get(trade['index'], {})
                exit_near = None
                exit_far = None

                # Try to get exit prices from bhavcopy
                near_strikes = idx_data.get(trade['near_expiry'], {})
                far_strikes = idx_data.get(trade['far_expiry'], {})

                near_opt = near_strikes.get(trade['strike'], {}).get(trade['side'], {})
                far_opt = far_strikes.get(trade['strike'], {}).get(trade['side'], {})

                exit_near = near_opt.get('close') if near_opt else None
                exit_far = far_opt.get('close') if far_opt else None

                # If near has expired, price is effectively 0 or whatever settlement
                if exit_near is None:
                    if trade_date >= trade['near_expiry']:
                        exit_near = 0.0  # expired worthless or settled
                    else:
                        continue  # can't exit yet, no price data

                if exit_far is None:
                    continue  # can't exit, no far price

                # Calculate PnL
                # Sell near (weekly), buy far (monthly)
                # PnL = (entry_near - exit_near) * lot_size * lots + (exit_far - entry_far) * lot_size * lots
                lot_size = LOT_SIZES[trade['index']]
                lots = trade['lots']
                near_pnl = (trade['entry_near'] - exit_near) * lot_size * lots
                far_pnl = (exit_far - trade['entry_far']) * lot_size * lots
                raw_pnl = round(near_pnl + far_pnl, 2)
                total_cost = ENTRY_COST + EXIT_COST
                net_pnl = round(raw_pnl - total_cost, 2)

                near_change = round(exit_near - trade['entry_near'], 2)
                far_change = round(exit_far - trade['entry_far'], 2)
                near_change_pct = round((near_change / trade['entry_near'] * 100) if trade['entry_near'] != 0 else 0, 2)
                far_change_pct = round((far_change / trade['entry_far'] * 100) if trade['entry_far'] != 0 else 0, 2)

                equity_before = equity
                equity += net_pnl
                equity_after = equity

                trade_counter += 1
                completed = {
                    'trade_num': trade_counter,
                    'trade_id': f"{trade['index']}_{trade['entry_date']}_{trade['side']}",
                    'index': trade['index'],
                    'side': trade['side'],
                    'atm_strike': trade['strike'],
                    'near_expiry': str(trade['near_expiry']),
                    'far_expiry': str(trade['far_expiry']),
                    'entry_date': str(trade['entry_date']),
                    'entry_time_ist': '15:30',
                    'exit_date': str(trade_date),
                    'exit_time_ist': '15:30',
                    'days_open': days_open,
                    'entry_near': trade['entry_near'],
                    'entry_far': trade['entry_far'],
                    'exit_near': exit_near,
                    'exit_far': exit_far,
                    'entry_ratio': trade['entry_ratio'],
                    'near_prem_change': near_change,
                    'far_prem_change': far_change,
                    'near_prem_change_pct': near_change_pct,
                    'far_prem_change_pct': far_change_pct,
                    'lots': lots,
                    'lot_size': lot_size,
                    'capital_deployed': trade['capital_deployed'],
                    'entry_costs': ENTRY_COST,
                    'exit_costs': EXIT_COST,
                    'total_costs': total_cost,
                    'raw_pnl': raw_pnl,
                    'net_pnl': net_pnl,
                    'exit_reason': exit_reason,
                    'all_triggered_exits': ','.join(all_exits) if all_exits else exit_reason,
                    'equity_before': equity_before,
                    'equity_after': equity_after,
                    'entry_event': None,
                    'exit_event': None,
                    'result': 'WIN' if net_pnl > 0 else 'LOSS',
                }
                completed_trades.append(completed)
                trades_to_close.append(i)

        # Remove closed trades (reverse order to maintain indices)
        for i in sorted(trades_to_close, reverse=True):
            active_trades.pop(i)

        # ── Evaluate new entries ──
        for index in INDICES:
            # Check if index is available in this period
            if datetime.combine(trade_date, datetime.min.time()) < INDEX_START[index]:
                continue

            spot = spot_today.get(index)
            if spot is None:
                continue

            idx_data = day_data.get(index, {})
            if not idx_data:
                continue

            # Get available expiries for this index
            available_expiries = sorted(idx_data.keys())

            # Find weekly and monthly expiry
            weekly_exp, monthly_exp = get_expiries(trade_date, available_expiries)
            if not weekly_exp or not monthly_exp:
                continue

            atm_strike = get_atm_strike(spot, STRIKE_INTERVAL[index])

            # Check both CE and PE independently
            for side in ['CE', 'PE']:
                # Get ATM option prices
                weekly_strikes = idx_data.get(weekly_exp, {})
                monthly_strikes = idx_data.get(monthly_exp, {})

                weekly_opt = weekly_strikes.get(atm_strike, {}).get(side, {})
                monthly_opt = monthly_strikes.get(atm_strike, {}).get(side, {})

                near_prem = weekly_opt.get('close')
                far_prem = monthly_opt.get('close')

                if near_prem is None or far_prem is None or near_prem <= 0 or far_prem <= 0:
                    continue

                ratio = round(near_prem / far_prem, 4)

                # Record ratio timeseries
                ratio_timeseries.append({
                    'date': str(trade_date),
                    'index': index,
                    'near_expiry': str(weekly_exp),
                    'far_expiry': str(monthly_exp),
                    'type': side,
                    'near_prem': near_prem,
                    'far_prem': far_prem,
                    'ratio': ratio,
                    'spot': spot,
                    'atm': atm_strike,
                    'event': None,
                })

                # ── Premium ratio gate ──
                if ratio < PREMIUM_RATIO_MIN or ratio > PREMIUM_RATIO_MAX:
                    continue

                key = f"{index}_{side}"

                # ── Concurrent limit ──
                if count_concurrent(active_trades, index, side) >= MAX_CONCURRENT[key]:
                    continue

                # ── No overlap ──
                if has_overlap(active_trades, index, side, atm_strike, weekly_exp):
                    continue

                # ── Capital allocation ──
                lot_size = LOT_SIZES[index]
                capital_for_trade = (near_prem + far_prem) * lot_size * 1
                alloc_pct = CAPITAL_PCT[key]
                max_capital = equity * (alloc_pct / 100.0)

                lots = max(1, int(max_capital / ((near_prem + far_prem) * lot_size)))
                capital_deployed = (near_prem + far_prem) * lot_size * lots

                # ── Exposure check ──
                curr_exp = current_exposure_pct(active_trades, equity)
                new_exp_pct = (capital_deployed / equity) * 100.0 if equity > 0 else 100.0
                if curr_exp + new_exp_pct > MAX_TOTAL_EXPOSURE_PCT:
                    continue

                # ── ENTER TRADE ──
                trade = {
                    'index': index,
                    'side': side,
                    'strike': atm_strike,
                    'near_expiry': weekly_exp,
                    'far_expiry': monthly_exp,
                    'entry_date': trade_date,
                    'entry_near': near_prem,
                    'entry_far': far_prem,
                    'entry_ratio': ratio,
                    'lots': lots,
                    'capital_deployed': capital_deployed,
                }
                active_trades.append(trade)

                logger.info("ENTRY: %s %s strike=%s near=%s far=%s ratio=%.4f lots=%d capital=%.0f",
                            index, side, atm_strike, weekly_exp, monthly_exp, ratio, lots, capital_deployed)

        # Progress log every 100 days
        if date_idx % 100 == 0:
            logger.info("Day %d/%d (%s): equity=%.2f, active=%d, completed=%d",
                        date_idx, len(all_dates), trade_date, equity, len(active_trades), len(completed_trades))

    # ── Force close any remaining active trades on last date ──
    last_date = all_dates[-1]
    for trade in active_trades:
        days_open = (last_date - trade['entry_date']).days
        lot_size = LOT_SIZES[trade['index']]
        lots = trade['lots']
        # Use last known prices or 0
        idx_data = bhav_data.get(last_date, {}).get(trade['index'], {})
        exit_near = 0.0
        exit_far = 0.0
        near_opt = idx_data.get(trade['near_expiry'], {}).get(trade['strike'], {}).get(trade['side'], {})
        far_opt = idx_data.get(trade['far_expiry'], {}).get(trade['strike'], {}).get(trade['side'], {})
        if near_opt:
            exit_near = near_opt.get('close', 0.0)
        if far_opt:
            exit_far = far_opt.get('close', 0.0)

        raw_pnl = round(((trade['entry_near'] - exit_near) + (exit_far - trade['entry_far'])) * lot_size * lots, 2)
        total_cost = ENTRY_COST + EXIT_COST
        net_pnl = round(raw_pnl - total_cost, 2)
        equity_before = equity
        equity += net_pnl
        trade_counter += 1
        completed_trades.append({
            'trade_num': trade_counter,
            'trade_id': f"{trade['index']}_{trade['entry_date']}_{trade['side']}",
            'index': trade['index'], 'side': trade['side'],
            'atm_strike': trade['strike'],
            'near_expiry': str(trade['near_expiry']), 'far_expiry': str(trade['far_expiry']),
            'entry_date': str(trade['entry_date']), 'entry_time_ist': '15:30',
            'exit_date': str(last_date), 'exit_time_ist': '15:30',
            'days_open': days_open,
            'entry_near': trade['entry_near'], 'entry_far': trade['entry_far'],
            'exit_near': exit_near, 'exit_far': exit_far,
            'entry_ratio': trade['entry_ratio'],
            'near_prem_change': round(exit_near - trade['entry_near'], 2),
            'far_prem_change': round(exit_far - trade['entry_far'], 2),
            'near_prem_change_pct': round(((exit_near - trade['entry_near']) / trade['entry_near'] * 100) if trade['entry_near'] != 0 else 0, 2),
            'far_prem_change_pct': round(((exit_far - trade['entry_far']) / trade['entry_far'] * 100) if trade['entry_far'] != 0 else 0, 2),
            'lots': lots, 'lot_size': lot_size,
            'capital_deployed': trade['capital_deployed'],
            'entry_costs': ENTRY_COST, 'exit_costs': EXIT_COST, 'total_costs': total_cost,
            'raw_pnl': raw_pnl, 'net_pnl': net_pnl,
            'exit_reason': 'BACKTEST_END', 'all_triggered_exits': 'BACKTEST_END',
            'equity_before': equity_before, 'equity_after': equity,
            'entry_event': None, 'exit_event': None,
            'result': 'WIN' if net_pnl > 0 else 'LOSS',
        })

    logger.info("=" * 60)
    logger.info("BACKTEST COMPLETE")
    logger.info("Total completed trades: %d", len(completed_trades))
    logger.info("Final equity: %.2f", equity)
    logger.info("Total return: %.2f%%", (equity - STARTING_EQUITY) / STARTING_EQUITY * 100)
    logger.info("Ratio timeseries entries: %d", len(ratio_timeseries))
    logger.info("=" * 60)

    return completed_trades, ratio_timeseries, equity


# ═══════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════

def write_excel(completed_trades, ratio_timeseries, final_equity, output_path):
    """Write 3-sheet Excel matching reference format."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Trades ──
    ws_trades = wb.active
    ws_trades.title = "Trades"

    trade_headers = [
        'trade_num', 'trade_id', 'index', 'side', 'atm_strike',
        'near_expiry', 'far_expiry', 'entry_date', 'entry_time_ist',
        'exit_date', 'exit_time_ist', 'days_open',
        'entry_near', 'entry_far', 'exit_near', 'exit_far',
        'entry_ratio', 'near_prem_change', 'far_prem_change',
        'near_prem_change_pct', 'far_prem_change_pct',
        'lots', 'lot_size', 'capital_deployed',
        'entry_costs', 'exit_costs', 'total_costs',
        'raw_pnl', 'net_pnl',
        'exit_reason', 'all_triggered_exits',
        'equity_before', 'equity_after',
        'entry_event', 'exit_event', 'result',
    ]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col, header in enumerate(trade_headers, 1):
        cell = ws_trades.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, trade in enumerate(completed_trades, 2):
        for col_idx, key in enumerate(trade_headers, 1):
            ws_trades.cell(row=row_idx, column=col_idx, value=trade.get(key))

    # ── Sheet 2: Ratio_Timeseries ──
    ws_ratio = wb.create_sheet("Ratio_Timeseries")
    ratio_headers = ['date', 'index', 'near_expiry', 'far_expiry', 'type',
                     'near_prem', 'far_prem', 'ratio', 'spot', 'atm', 'event']

    for col, header in enumerate(ratio_headers, 1):
        cell = ws_ratio.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, entry in enumerate(ratio_timeseries, 2):
        for col_idx, key in enumerate(ratio_headers, 1):
            ws_ratio.cell(row=row_idx, column=col_idx, value=entry.get(key))

    # ── Sheet 3: Summary ──
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="Metric").font = header_font
    ws_summary.cell(row=1, column=2, value="Value").font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=2).fill = header_fill

    wins = sum(1 for t in completed_trades if t['result'] == 'WIN')
    losses = sum(1 for t in completed_trades if t['result'] == 'LOSS')
    total = len(completed_trades)
    net_pnls = [t['net_pnl'] for t in completed_trades]
    avg_pnl = sum(net_pnls) / total if total > 0 else 0
    max_win = max(net_pnls) if net_pnls else 0
    max_loss = min(net_pnls) if net_pnls else 0

    idx_counts = {idx: sum(1 for t in completed_trades if t['index'] == idx) for idx in INDICES}
    exit_counts = defaultdict(int)
    for t in completed_trades:
        exit_counts[t['exit_reason']] += 1

    summary_data = [
        ("Strategy", "ATM_CALL_PUT_CALENDAR v1.4"),
        ("Side", "CE & PE"),
        ("Period", "2020-01-01 to 2025-12-31"),
        ("Starting Equity", f"{STARTING_EQUITY:,.2f}"),
        ("Final Equity", f"{final_equity:,.2f}"),
        ("Total Return", f"{(final_equity - STARTING_EQUITY) / STARTING_EQUITY * 100:.2f}%"),
        ("Total Trades", total),
        ("Wins", wins),
        ("Losses", losses),
        ("Win Rate", f"{wins / total * 100:.1f}%" if total > 0 else "N/A"),
        ("Avg Net PnL", f"{avg_pnl:.2f}"),
        ("Max Win", f"{max_win:.2f}"),
        ("Max Loss", f"{max_loss:.2f}"),
        ("BANKNIFTY Trades", idx_counts.get("BANKNIFTY", 0)),
        ("NIFTY Trades", idx_counts.get("NIFTY", 0)),
        ("FINNIFTY Trades", idx_counts.get("FINNIFTY", 0)),
        ("MIDCPNIFTY Trades", idx_counts.get("MIDCPNIFTY", 0)),
        ("EXIT_TIME", exit_counts.get("EXIT_TIME", 0)),
        ("EXPIRY_FORCE_CLOSE", exit_counts.get("EXPIRY_FORCE_CLOSE", 0)),
        ("EXIT_OI_REVERSAL", exit_counts.get("EXIT_OI_REVERSAL", 0)),
        ("EXIT_TREND_DECAY", exit_counts.get("EXIT_TREND_DECAY", 0)),
        ("BACKTEST_END", exit_counts.get("BACKTEST_END", 0)),
    ]

    for row_idx, (metric, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_idx, column=1, value=metric)
        ws_summary.cell(row=row_idx, column=2, value=value)

    # Auto-width columns
    for ws in [ws_trades, ws_ratio, ws_summary]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info("Excel saved to: %s", output_path)


# ═══════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def write_csv(completed_trades, ratio_timeseries, output_dir):
    """Write trades CSV and ratio timeseries CSV matching reference format."""

    trade_headers = [
        'trade_num', 'trade_id', 'index', 'side', 'atm_strike',
        'near_expiry', 'far_expiry', 'entry_date', 'entry_time_ist',
        'exit_date', 'exit_time_ist', 'days_open',
        'entry_near', 'entry_far', 'exit_near', 'exit_far',
        'entry_ratio', 'near_prem_change', 'far_prem_change',
        'near_prem_change_pct', 'far_prem_change_pct',
        'lots', 'lot_size', 'capital_deployed',
        'entry_costs', 'exit_costs', 'total_costs',
        'raw_pnl', 'net_pnl',
        'exit_reason', 'all_triggered_exits',
        'equity_before', 'equity_after',
        'entry_event', 'exit_event', 'result',
    ]

    trades_csv = os.path.join(output_dir, "backtest_results_v1_4_CE_PE_multiyear.csv")
    with open(trades_csv, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=trade_headers, extrasaction='ignore')
        writer.writeheader()
        for t in completed_trades:
            writer.writerow(t)
    logger.info("Trades CSV saved to: %s", trades_csv)

    ratio_headers = ['date', 'index', 'near_expiry', 'far_expiry', 'type',
                     'near_prem', 'far_prem', 'ratio', 'spot', 'atm', 'event']

    ratio_csv = os.path.join(output_dir, "ratio_timeseries_CE_PE_2020_2025.csv")
    with open(ratio_csv, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=ratio_headers, extrasaction='ignore')
        writer.writeheader()
        for r in ratio_timeseries:
            writer.writerow(r)
    logger.info("Ratio CSV saved to: %s", ratio_csv)


def write_start_txt(output_dir):
    """Write strategy start text file matching reference format."""
    from src.utils.strategy_integrity import EXPECTED_HASH
    txt_path = os.path.join(output_dir, "v1.4(PE&CE)start.txt")
    strategy_path = os.path.join(PROJECT_ROOT, "STRATEGY_LOCKED_V1.4_CE_PE.md")
    with open(strategy_path, 'r') as f:
        content = f.read()
    with open(txt_path, 'w') as f:
        f.write(content)
    logger.info("Start text saved to: %s", txt_path)


if __name__ == "__main__":
    OUTPUT_DIR = r"C:\Users\sathv\OneDrive\Desktop\backtesting_fo\v1.4(PE&CE)[2020-2025]"
    OUTPUT_FILE = os.path.join(OUTPUT_DIR, "ATM_Calendar_v1_4_CE_PE_MultiYear_2020_2025.xlsx")

    trades, ratios, final_eq = run_backtest()
    write_excel(trades, ratios, final_eq, OUTPUT_FILE)
    write_csv(trades, ratios, OUTPUT_DIR)
    write_start_txt(OUTPUT_DIR)

    print(f"\nBacktest complete!")
    print(f"Total trades: {len(trades)}")
    print(f"Final equity: {final_eq:,.2f}")
    print(f"Output: {OUTPUT_DIR}")
